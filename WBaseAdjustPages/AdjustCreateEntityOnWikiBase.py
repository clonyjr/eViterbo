# load the necessary libraries
import csv
import openpyxl
import sys
import requests
import logging

from wikibase_api import Wikibase
import chardet
import pandas as pd
from wikidataintegrator.wdi_config import config
from wikidataintegrator import wdi_login, wdi_core

PATH_EXCEL_PROPERTIES = "/Users/clonyjr/Library/Mobile Documents/com~apple~CloudDocs/Aveiro/UA/CLONY/TechnetEmpire/WBaseAdjustPages/Propriedades.xlsx"
datasetpath = "/Users/clonyjr/Library/Mobile Documents/com~apple~CloudDocs/Aveiro/UA/CLONY/TechnetEmpire/" \
              "WBaseAdjustPages/"
coord_file = "coordenadasWikibase-May2020.csv"
coord_file2 = "coordfinal.csv"
instituicao_file = "instituicaoWikibase-May2020.csv"
pessoa_file = "criar_entidades_vazias _teste.csv" #"Query Results.csv"#"pessoas_wikibase_restante2_julho_2020.csv"
URL = "http://ieeta-eviterbo.web.ua.pt/api.php"
URL_EVITERBO = "http://ieeta-eviterbo.web.ua.pt/index.php/"
object_without_properties = {'value': None, 'snak_type': 'somevalue'}
fmtstr = " Log: %(asctime)s: %(levelname)s: %(funcName)s Line: %(lineno)d - %(message)s"
datestr = "%d/%m/%Y %I:%M:%S %p "
#basic logging config
logging.basicConfig(
    filename="wikibase_eviterbo_log_output.log",
    level=logging.DEBUG,
    filemode="w",
    format=fmtstr,
    datefmt=datestr,
)

# WikidataIntegrator Configuration
config['MEDIAWIKI_API_URL'] = "http://ieeta-eviterbo.web.ua.pt/api.php"
config['WIKIBASE_URL'] = 'http://ieeta-eviterbo.web.ua.pt'
config['CONCEPT_BASE_URI'] = 'http://ieeta-eviterbo.web.ua.pt/entity/'
config['CALENDAR_MODEL_QID'] = 'http://www.wikidata.org/entity/Q1985727'
config['COORDINATE_GLOBE_QID'] = 'http://www.wikidata.org/entity/Q2'
columns_Pessoas = ['_pageName','_pageID','nome_completo','nome_outras_grafias','pai','mãe','cônjuge','filhos','irmãos'
    ,'data_nascimento','local_nascimento','data_morte','local_morte','residência1','data_inicio_residência1'
    ,'data_fim_residência1','residência2','data_inicio_residência2','data_fim_residência2','residência3'
    ,'data_inicio_residência3','data_fim_residência3','residência4','data_inicio_residência4'
    ,'data_fim_residência4','residência5','data_inicio_residência5','data_fim_residência5','residência6'
    ,'data_inicio_residência6','data_fim_residência6','residência7','data_inicio_residência7'
    ,'data_fim_residência7','residência8','data_inicio_residência8','data_fim_residência8','residência9'
    ,'data_inicio_residência9','data_fim_residência9','religião','sexo','local_enterramento'
    ,'Formação_1','Instituição_de_Formação_1','data_inicio_formação_1','data_fim_formação_1','Formação_2'
    ,'Instituição_de_Formação_2','data_inicio_formação_2','data_fim_formação_2','Formação_3'
    ,'Instituição_de_Formação_3','data_inicio_formação_3','data_fim_formação_3','Posto_1','Arma_1'
    ,'data_posto_1','data_fim_posto_1','Posto_2','Arma_2','data_posto_2','data_fim_posto_2','Posto_3','Arma_3'
    ,'data_posto_3','data_fim_posto_3','Posto_4','Arma_4','data_posto_4','data_fim_posto_4','Posto_5','Arma_5'
    ,'data_posto_5','data_fim_posto_5','Posto_6','Arma_6','data_posto_6','data_fim_posto_6','Posto_7','Arma_7'
    ,'data_posto_7','data_fim_posto_7','Posto_8','Arma_8','data_posto_8','data_fim_posto_8','Cargo_1'
    ,'Instituição_Cargo_1','data_inicio_cargo_1','data_fim_cargo_1','Cargo_2','Instituição_Cargo_2'
    ,'data_inicio_cargo_2','data_fim_cargo_2','Cargo_3','Instituição_Cargo_3','data_inicio_cargo_3'
    ,'data_fim_cargo_3','Cargo_4','Instituição_Cargo_4','data_inicio_cargo_4','data_fim_cargo_4','Cargo_5'
    ,'Instituição_Cargo_5','data_inicio_cargo_5','data_fim_cargo_5','Cargo_6','Instituição_Cargo_6'
    ,'data_inicio_cargo_6','data_fim_cargo_6','Cargo_7','Instituição_Cargo_7','data_inicio_cargo_7'
    ,'data_fim_cargo_7','Cargo_8','Instituição_Cargo_8','data_inicio_cargo_8','data_fim_cargo_8','Cargo_9'
    ,'Instituição_Cargo_9','data_inicio_cargo_9','data_fim_cargo_9','Actividade_1','data_inicio_actividade_1'
    ,'data_fim_actividade_1','Local_de_Actividade_1','Actividade_2','data_inicio_actividade_2'
    ,'data_fim_actividade_2','Local_de_Actividade_2','Actividade_3','data_inicio_actividade_3'
    ,'data_fim_actividade_3','Local_de_Actividade_3','Actividade_4','data_inicio_actividade_4'
    ,'data_fim_actividade_4','Local_de_Actividade_4','Actividade_5','data_inicio_actividade_5'
    ,'data_fim_actividade_5','Local_de_Actividade_5','Actividade_6','data_inicio_actividade_6'
    ,'data_fim_actividade_6','Local_de_Actividade_6','Actividade_7','data_inicio_actividade_7'
    ,'data_fim_actividade_7','Local_de_Actividade_7','Actividade_8','data_inicio_actividade_8'
    ,'data_fim_actividade_8','Local_de_Actividade_8','Actividade_9','data_inicio_actividade_9'
    ,'data_fim_actividade_9','Local_de_Actividade_9']
column_formacao = ['Formação_1','Formação_2','Formação_3']
column_residencia = ['residência1','residência2','residência3','residência4',
                     'residência5','residência6','residência7','residência8','residência9']
column_cargo =  ['Cargo_1','Cargo_2','Cargo_3','Cargo_4','Cargo_5','Cargo_6','Cargo_7','Cargo_8','Cargo_9']
column_inst_cargo = ['Instituição_Cargo_1','Instituição_Cargo_2','Instituição_Cargo_3','Instituição_Cargo_4',
                     'Instituição_Cargo_5','Instituição_Cargo_6','Instituição_Cargo_7','Instituição_Cargo_8',
                     'Instituição_Cargo_9']
column_posto =  ['Posto_1','Posto_2','Posto_3','Posto_4','Posto_5','Posto_6','Posto_7','Posto_8']
column_arma =  ['Arma_1','Arma_2','Arma_3','Arma_4','Arma_5','Arma_6','Arma_7','Arma_8']
column_actividade =  ['Actividade_1','Actividade_2','Actividade_3','Actividade_4','Actividade_5','Actividade_6',
                      'Actividade_7','Actividade_8','Actividade_9']
column_loc_actividade =  ['Local_de_Actividade_1','Local_de_Actividade_2','Local_de_Actividade_3',
                          'Local_de_Actividade_4','Local_de_Actividade_5','Local_de_Actividade_6',
                          'Local_de_Actividade_7','Local_de_Actividade_8','Local_de_Actividade_9']

datas_qualifiers={'residência':['data_inicio_residência','data_fim_residência'],
                  'formação':['data_inicio_formação_','data_fim_formação_'],
                  'posto':['data_posto_','data_fim_posto_'],
                  'cargo':['data_inicio_cargo_','data_fim_cargo_'],
                  'actividade':['data_inicio_actividade_','data_fim_actividade_']}
dfWikibaseItems = pd.DataFrame
property_with_no_value = '''None, snak_type='somevalue'''
religioes = {'Católica': 'Q8'
    ,'Cristã Ortodoxa': 'Q69'
    ,'Protestante': 'Q68'
    ,'Anglicana': 'Q67'
    ,'Hindu': 'Q66'
    ,'Budista': 'Q65'
    ,'Islmâmica': 'Q64'
    ,'Judaica': 'Q63'}

class WikibaseWriter(object):
    def __init__(self):
        self.login_credentials = {
            "bot_username": "Operador@evitWikiBase",
            "bot_password": "kf2eei97qac85brujd5hii0fae0ok3vk",
        }
        self.login_instance = wdi_login.WDLogin(user= "Operador@evitWikiBase", pwd="kf2eei97qac85brujd5hii0fae0ok3vk")
        excelfile = openpyxl.load_workbook(PATH_EXCEL_PROPERTIES)
        self.ws = excelfile.active
        self.maxRows = self.ws.max_row
        self.maxCol = self.ws.max_column

    def initiate(self):
        r_file = open(datasetpath+pessoa_file, 'rb').read()
        result = chardet.detect(r_file)
        file_encoding = result['encoding']
        df_pessoas = pd.read_csv(datasetpath+pessoa_file
                                 ,encoding=file_encoding
                                 ,header=0
                                 ,usecols=columns_Pessoas)
        df_pessoas = df_pessoas.reindex(columns=df_pessoas.columns.tolist())
        df_pessoas_nomes = df_pessoas['_pageName'].replace(' ', '_', regex=True)
        df_pessoas = df_pessoas.replace(['\[','\]'], ['',''], regex=True)
        df_pessoas = df_pessoas.fillna(0)
        return df_pessoas

    def get_qualifiers(self,prop_qualifier,data_ini,data_fim):
        qualifiers_result = []
        if 'P9' == prop_qualifier and data_ini is not None:
            if data_ini != '0' and data_fim != '0':
                qualifiers_result.append(wdi_core.WDTime(time=self.parse_To_Wikibase_Time(data_ini), prop_nr=prop_qualifier, is_qualifier=True))
                qualifiers_result.append(wdi_core.WDTime(time=self.parse_To_Wikibase_Time(data_fim),prop_nr='P10', is_qualifier=True))
            elif data_ini != '0' and data_fim == '0':
                qualifiers_result.append(wdi_core.WDTime(time=self.parse_To_Wikibase_Time(data_ini), prop_nr=prop_qualifier, is_qualifier=True))
                qualifiers_result.append(wdi_core.WDTime(prop_nr='P10',time=None,snak_type='somevalue', is_qualifier=True))
            elif data_ini == '0' and data_fim != '0':
                qualifiers_result.append(wdi_core.WDTime(prop_nr='P9',time=None,snak_type='somevalue', is_qualifier=True))
                qualifiers_result.append(wdi_core.WDTime(time=self.parse_To_Wikibase_Time(data_fim),prop_nr='P10', is_qualifier=True))
            elif data_ini == '0' and data_fim == '0':
                qualifiers_result.append(wdi_core.WDTime(prop_nr='P9',time=None,snak_type='somevalue', is_qualifier=True))
                qualifiers_result.append(wdi_core.WDTime(prop_nr='P10',time=None,snak_type='somevalue', is_qualifier=True))
        return qualifiers_result


    def add_properties_and_qualifiers(self, item_id,df_complete,idx_df, prop, prop_qualifier):
        inst_name = df_complete['_pageName'][idx_df]
        dt_qualif_loc_1 = []
        dt_qualif_loc_2 = []
        dt_qualif_loc_3 = []
        dt_qualif_loc_4 = []
        dt_qualif_loc_5 = []
        dt_qualif_loc_6 = []
        dt_qualif_loc_7 = []
        dt_qualif_loc_8 = []
        data = []
        j=1
        max_size = 1
        while j <= max_size:
            column_name = 'Localização' + str(j)
            localizacao_dt_ini_x = 'Localização_data_início' + str(j)
            localizacao_dt_fim_x = 'Localização_data_fim' + str(j)
            data_ini = df_complete[localizacao_dt_ini_x][idx_df]
            data_fim = df_complete[localizacao_dt_fim_x][idx_df]
            print("Updating  {}  ...".format(inst_name))
            Q_instituicao = self.get_Wikibase_Qid(inst_name,'Instituição')
            print("Inserting location {} to item {},  ".format(column_name, Q_instituicao))
            prop_QID = []
            prop_QID = self.get_Wikibase_Qid(df_complete[column_name][idx_df])
            data_ini = self.adjust_Time_Value(data_ini)
            data_fim = self.adjust_Time_Value(data_fim)
            if j==1:
                dt_qualif_loc_1 = self.get_qualifiers(prop_qualifier,data_ini,data_fim)
            elif j==2:
                dt_qualif_loc_2 = self.get_qualifiers(prop_qualifier,data_ini,data_fim)
            elif j==3:
                dt_qualif_loc_3 = self.get_qualifiers(prop_qualifier,data_ini,data_fim)
            elif j==4:
                dt_qualif_loc_4 = self.get_qualifiers(prop_qualifier,data_ini,data_fim)
            elif j==5:
                dt_qualif_loc_5 = self.get_qualifiers(prop_qualifier,data_ini,data_fim)
            elif j==6:
                dt_qualif_loc_6 = self.get_qualifiers(prop_qualifier,data_ini,data_fim)
            elif j==7:
                dt_qualif_loc_7 = self.get_qualifiers(prop_qualifier,data_ini,data_fim)
            elif j==8:
                dt_qualif_loc_8 = self.get_qualifiers(prop_qualifier,data_ini,data_fim)

            j+=1
        data.append(wdi_core.WDItemID(value=prop_QID['value'], snak_type=prop_QID['snak_type'],prop_nr=prop, qualifiers=dt_qualif_loc_1))
        #data.append(wdi_core.WDItemID(value=prop_QID['value'], snak_type=prop_QID['snak_type'],prop_nr=prop, qualifiers=dt_qualif_loc_2))
        #data.append(wdi_core.WDItemID(value=prop_QID['value'], snak_type=prop_QID['snak_type'],prop_nr=prop, qualifiers=dt_qualif_loc_3))
        #data.append(wdi_core.WDItemID(value=prop_QID['value'], snak_type=prop_QID['snak_type'],prop_nr=prop, qualifiers=dt_qualif_loc_4))
        #data.append(wdi_core.WDItemID(value=prop_QID['value'], snak_type=prop_QID['snak_type'],prop_nr=prop, qualifiers=dt_qualif_loc_5))
        #data.append(wdi_core.WDItemID(value=prop_QID['value'], snak_type=prop_QID['snak_type'],prop_nr=prop, qualifiers=dt_qualif_loc_6))
        #data.append(wdi_core.WDItemID(value=prop_QID['value'], snak_type=prop_QID['snak_type'],prop_nr=prop, qualifiers=dt_qualif_loc_7))
        #data.append(wdi_core.WDItemID(value=prop_QID['value'], snak_type=prop_QID['snak_type'],prop_nr=prop, qualifiers=dt_qualif_loc_8))

        wd_item = wdi_core.WDItemEngine(wd_item_id=item_id['value'], data=data)
        wd_item.write(self.login_instance,edit_summary='Adicionando data de inicio e fim das localizações')
        return wd_item.entity_metadata.get('id')

    def create_new_item_on_wiki(self, new_item, description_item):
        result = None
        wd_item = wdi_core.WDItemEngine() # create new item
        wd_item.set_label(label=new_item,lang='pt')
        wd_item.set_description(description=description_item,lang='pt')
        wd_item.write(self.login_instance)
        result = wd_item.entity_metadata.get('id') #retrieve item id
        return result

    def get_Properties_Values(self, value):
        value_to_return = {'value': None, 'snak_type': 'somevalue'}
        if(value != 0):
            value_to_return = {'value': value, 'snak_type': 'value'}
        return value_to_return

    def adjust_Time_Value(self, value):
        value_to_return = str(value)
        if value_to_return.count('.') > 0:
            value_to_return = value_to_return[:-1].replace('.','')
            print(value_to_return)
            print(value_to_return[:-1])
        return value_to_return

    def parse_To_Wikibase_Time(self, stringTime):
        result = 'invalid'
        year = '+'
        day = ''
        month = ''
        zeromonthday='-00-00'
        zeroday = '-00'
        calendar = {'janeiro': '01'
            ,'fevereiro': '02'
            ,'março': '03'
            ,'abril': '04'
            ,'maio': '05'
            ,'junho': '06'
            ,'julho': '07'
            ,'agosto': '08'
            ,'setembro': '09'
            ,'outubro': '10'
            ,'novembro': '11'
            ,'dezembro': '12'
                 }
        hour = 'T00:00:00Z'

        try:
            stringTime = stringTime.replace(' ','-')
            if(stringTime.count('-') == 4):
                if('-' in stringTime[0:2]):
                    day = '-0' + stringTime[0:1]
                    month = stringTime[5:-8]
                else:
                    day = '-' + stringTime[0:2]
                    month = stringTime[6:-8]
                year = year + stringTime[-4:]+'-'
                result = year+calendar[month]+day+hour
            elif (stringTime.count('-') == 3):
                if('-' in stringTime[0:2]):
                    day = '-0' + stringTime[0:1]
                    month = stringTime[5:-5]
                else:
                    day = '-' + stringTime[0:2]
                    month = stringTime[6:-5]
                year = year + stringTime[-4:]
                result = year+'-'+calendar[month]+day+hour
            elif (stringTime.count('-') == 2):
                day = zeroday
                year = year + stringTime[-4:]
                month = stringTime[0:-8]
                result = year+'-'+calendar[month]+day+hour
            elif (stringTime.count('-') == 0):
                year = year+stringTime[:4]
                result = year+zeromonthday+hour
        except:
             sys.exit

        return result

    def get_Wikibase_Qid(self, item_name, description=None):
        if(item_name == 'Angra do Heroísmo, Açores, Portugal'):
            result = {'value': 'Q464'
                ,'snak_type': 'value'}
        else:
            result = {'value': None
            ,'snak_type': 'somevalue'
                        } #'''None, snak_type='somevalue'''
        if(item_name != 0):
            search_items = wdi_core.WDItemEngine.get_wd_search_results(
                search_string=item_name[:-1], language='pt')
            if len(search_items) > 0:
                if(description == 'Instituição'):
                    str_test = search_items[0]
                    str_test = str_test[1:]
                    if 478 <= int(str_test) <= 606:
                        result={'value':search_items[0]
                            ,'snak_type':'value'}
                    else: result={'value':search_items[1]
                        ,'snak_type':'value'}
                else: result={'value':search_items[0]
                    ,'snak_type':'value'}
        else: result = {'value': None
            ,'snak_type': 'somevalue'}
        return result

    def check_empty_value(self,dataframe_to_check,index_to_check,column_name):
        localizacao = []
        for index_to_check in range(len(column_name)):
            if(dataframe_to_check.iloc[index_to_check][column_name[7]]!=0):
                localizacao.append('L1:' + dataframe_to_check.iloc[index_to_check][column_name[7]])
            if(dataframe_to_check.iloc[index_to_check][column_name[10]]!=0):
                localizacao.append('L2:' + dataframe_to_check.iloc[index_to_check][column_name[10]])
            if(dataframe_to_check.iloc[index_to_check][column_name[13]]!=0):
                localizacao.append('L3:' + dataframe_to_check.iloc[index_to_check][column_name[13]])
            if(dataframe_to_check.iloc[index_to_check][column_name[16]]!=0):
                localizacao.append('L4:' + dataframe_to_check.iloc[index_to_check][column_name[16]])
            if(dataframe_to_check.iloc[index_to_check][column_name[19]]!=0):
                localizacao.append('L5:' + dataframe_to_check.iloc[index_to_check][column_name[19]])
            if(dataframe_to_check.iloc[index_to_check][column_name[22]]!=0):
                localizacao.append('L6:' + dataframe_to_check.iloc[index_to_check][column_name[22]])
            if(dataframe_to_check.iloc[index_to_check][column_name[25]]!=0):
                localizacao.append('L7:' + dataframe_to_check.iloc[index_to_check][column_name[25]])
            if(dataframe_to_check.iloc[index_to_check][column_name[28]]!=0):
                localizacao.append('L8:' + dataframe_to_check.iloc[index_to_check][column_name[28]])

        #qualifiers = []
        #qualifiers.append(wdi_core.WDTime(prop_nr='P9', time='+1837-03-01T00:00:00Z', is_qualifier=True))
        #qualifiers.append(wdi_core.WDTime(prop_nr='P10', time='+1911-00-00T00:00:00Z', is_qualifier=True))
        #wd_item = wdi_core.WDItemEngine(wd_item_id='Q48', data=[wdi_core.WDString(prop_nr='P67', value='SIM', qualifiers=qualifiers)])
        #wd_item = wdi_core.WDItemEngine()
        # print("=== PRINTING THE LOCALIZATION ARRAY ===")
        # for x in range(len(localizacao)):
        #     print(localizacao[x])

        return localizacao

    '''=== POPULATE INSTITUICOES === '''
    def create_instituicoes(self):
        columns_Instituicao = ['_pageName','outras_denominações','tipo','data_fundação','data_extinção'
            ,'paralisação_início','paralisação_fim','Localização1','Localização_data_início1'
            ,'Localização_data_fim1','Localização2','Localização_data_início2','Localização_data_fim2'
            ,'Localização3','Localização_data_início3','Localização_data_fim3','Localização4'
            ,'Localização_data_início4','Localização_data_fim4','Localização5','Localização_data_início5'
            ,'Localização_data_fim5','Localização6','Localização_data_início6','Localização_data_fim6','Localização7'
            ,'Localização_data_início7','Localização_data_fim7','Localização8','Localização_data_início8'
            ,'Localização_data_fim8','antecessora','sucessora']
        for i in range(len(columns_Instituicao)):
            print(columns_Instituicao[i])
        r_file = open(datasetpath+coord_file, 'rb').read()
        result = chardet.detect(r_file)
        file_encoding = result['encoding']
        dfInstituicao = pd.read_csv(datasetpath+instituicao_file
                                    ,encoding=file_encoding
                                    ,header=0
                                    ,usecols=columns_Instituicao)
        dfInstituicao = dfInstituicao.reindex(columns=dfInstituicao.columns.tolist())
        dfInstNames = dfInstituicao['_pageName'].replace(' ', '_', regex=True)
        dfInstituicao = dfInstituicao.replace(['\[','\]'], ['',''], regex=True)
        dfInstituicao = dfInstituicao.fillna(0)
        instituicao = {}
        print(dfInstNames.head(10))
        for i in range(len(dfInstituicao)):
            inst_name = dfInstituicao['_pageName'][i]
            print(inst_name.replace(' ','_'))
            url_coord=URL_EVITERBO + dfInstituicao['_pageName'].replace(' ', '_', regex=True)[i]
            print(url_coord)
            outras_denominações = self.get_Properties_Values(dfInstituicao['outras_denominações'][i])
            tipo = self.get_Properties_Values(dfInstituicao['tipo'][i])
            local1 = self.get_Wikibase_Qid(dfInstituicao['Localização1'][i])
            local2 = self.get_Wikibase_Qid(dfInstituicao['Localização2'][i])
            local3 = self.get_Wikibase_Qid(dfInstituicao['Localização3'][i])
            local4 = self.get_Wikibase_Qid(dfInstituicao['Localização4'][i])
            local5 = self.get_Wikibase_Qid(dfInstituicao['Localização5'][i])
            local6 = self.get_Wikibase_Qid(dfInstituicao['Localização6'][i])
            local7 = self.get_Wikibase_Qid(dfInstituicao['Localização7'][i])
            local8 = self.get_Wikibase_Qid(dfInstituicao['Localização8'][i])
            antecessora = self.get_Wikibase_Qid(dfInstituicao['antecessora'][i])
            sucessora = self.get_Wikibase_Qid(dfInstituicao['sucessora'][i])
            data=[wdi_core.WDUrl(prop_nr='P30',value=url_coord)
                ,wdi_core.WDGlobeCoordinate(prop_nr='P29',latitude=None,longitude=None,precision=None,snak_type='somevalue')
                ,wdi_core.WDString(prop_nr='P68', value=outras_denominações['value'],snak_type=outras_denominações['snak_type'])
                ,wdi_core.WDString(prop_nr='P69', value=None, snak_type='somevalue')
                ,wdi_core.WDItemID(prop_nr='P35', value=local1['value'],snak_type=local1['snak_type'])
                ,wdi_core.WDItemID(prop_nr='P35', value=local2['value'],snak_type=local2['snak_type'])
                ,wdi_core.WDItemID(prop_nr='P35', value=local3['value'],snak_type=local3['snak_type'])
                ,wdi_core.WDItemID(prop_nr='P35', value=local4['value'],snak_type=local4['snak_type'])
                ,wdi_core.WDItemID(prop_nr='P35', value=local5['value'],snak_type=local5['snak_type'])
                ,wdi_core.WDItemID(prop_nr='P35', value=local6['value'],snak_type=local6['snak_type'])
                ,wdi_core.WDItemID(prop_nr='P35', value=local7['value'],snak_type=local7['snak_type'])
                ,wdi_core.WDItemID(prop_nr='P35', value=local8['value'],snak_type=local8['snak_type'])
                ,wdi_core.WDItemID(prop_nr='P60', value=antecessora['value'],snak_type=antecessora['snak_type'])
                ,wdi_core.WDItemID(prop_nr='P61', value=sucessora['value'],snak_type=sucessora['snak_type'])
                ,wdi_core.WDString(prop_nr='P36', value=tipo['value'],snak_type=tipo['snak_type'])
                ,wdi_core.WDString(prop_nr='P45', value=dfInstituicao['_pageName'][i])]
            if(dfInstituicao['paralisação_início'][i] != 0):
                data.append(wdi_core.WDTime(prop_nr='P66', time=self.parse_To_Wikibase_Time(dfInstituicao['paralisação_início'][i])))
            else: data.append(wdi_core.WDTime(prop_nr='P66', time=None, snak_type='somevalue'))
            if(dfInstituicao['paralisação_fim'][i] != 0):
                data.append(wdi_core.WDTime(prop_nr='P70', time=self.parse_To_Wikibase_Time(dfInstituicao['paralisação_fim'][i])))
            else: data.append(wdi_core.WDTime(prop_nr='P70', time=None, snak_type='somevalue'))
            if(dfInstituicao['data_fundação'][i] != 0):
                data.append(wdi_core.WDTime(prop_nr='P62', time=self.parse_To_Wikibase_Time(dfInstituicao['data_fundação'][i])))
            else: data.append(wdi_core.WDTime(prop_nr='P62', time=None, snak_type='somevalue'))
            if(dfInstituicao['data_extinção'][i] != 0):
                data.append(wdi_core.WDTime(prop_nr='P63', time=self.parse_To_Wikibase_Time(dfInstituicao['data_extinção'][i])))
            else: data.append(wdi_core.WDTime(prop_nr='P63', time=None, snak_type='somevalue'))
            wd_item = wdi_core.WDItemEngine(data=data) # create new item
            wd_item.set_label(label=dfInstituicao['_pageName'][i],lang='pt')
            wd_item.set_description(description='Instituição',lang='pt')
            wd_item.write(self.login_instance)
            item_id = wd_item.entity_metadata.get('id') #retrieve item id

            instituicao = {
                'nome': dfInstituicao['_pageName'],
                'urleviterbo': URL_EVITERBO + dfInstituicao['_pageName'].replace(' ', '_', regex=True),
                'outras_denominações': dfInstituicao['outras_denominações'],
                'sigla': 'value = None, snak_type = somevalue',
                'local_inst': dfInstituicao['Localização1'],
                'tipo_inst': dfInstituicao['tipo'],
                'antecessora':dfInstituicao['antecessora'],
                'sucessora':dfInstituicao['sucessora'],
                'data_fundação':dfInstituicao['data_fundação'],
                'data_extinção':dfInstituicao['data_extinção'],
                'paralisação_início':dfInstituicao['paralisação_início'],
                'paralisação_fim':dfInstituicao['paralisação_fim'],
                'item_id': item_id
            }
            dfWikibaseItems = pd.DataFrame(instituicao,columns=['nome'
                ,'urleviterbo'
                ,'outras_denominações'
                ,'sigla'
                ,'local_inst'
                ,'tipo_inst'
                ,'antecessora'
                ,'sucessora'
                ,'data_fundação'
                ,'data_extinção'
                ,'paralisação_início'
                ,'paralisação_fim'
                ,'item_id'])
            print(dfWikibaseItems)
            dfWikibaseItems.to_csv(datasetpath+'instituicao_wikibase_item_Teste.csv', index = True, header=True)

    '''=== POPULATE COORDENADAS ==='''
    def create_coordenadas(self):
        dfWikibaseItems = pd.DataFrame
        #wd_item = wdi_core.WDItemEngine(data=data) # create new item
        #wd_item.entity_metadata.get('id') #retriev item id
        columns_Coordenada = ['_pageName','1','5']
        r_file = open(datasetpath+coord_file, 'rb').read()
        result = chardet.detect(r_file)
        file_encoding = result['encoding']

        dfCoordenadas = pd.read_csv(datasetpath+coord_file2
                                    ,encoding=file_encoding
                                    ,header=0
                                    #,skiprows=[0,124]
                                    ,nrows=22
                                    ,usecols=columns_Coordenada)


        dfCoordenadas = dfCoordenadas.reindex(columns=dfCoordenadas.columns.tolist())
        dataCoordNames = dfCoordenadas['_pageName'].replace(' ', '_', regex=True)
        #dataTemp = dataTemp.replace(['\[','\]'], ['',''], regex=True)
        coordenadas = {}
        print(dataCoordNames.head(10))
        #frame = pd.DataFrame(dataCoordNames['_pageName'])
        #print(dataCoordNames.values)
        for i in range(len(dfCoordenadas)):
            coord_name = dfCoordenadas['_pageName'][i]
            print(coord_name.replace(' ','_'))
            url_coord=URL_EVITERBO + dfCoordenadas['_pageName'].replace(' ', '_', regex=True)[i]
            print(url_coord)
            lat = dfCoordenadas['1'][i]
            long = dfCoordenadas['5'][i]
            data=[wdi_core.WDUrl(prop_nr='P30',value=url_coord)
                ,wdi_core.WDGlobeCoordinate(prop_nr='P29', latitude=lat,longitude=long,precision=0.0001)
                ,wdi_core.WDString(prop_nr='P45', value=dfCoordenadas['_pageName'][i])]
            wd_item = wdi_core.WDItemEngine(data=data) # create new item
            wd_item.set_label(label=dfCoordenadas['_pageName'][i],lang='pt')
            item_id = wd_item.entity_metadata.get('id') #retrieve item id
            wd_item.write(self.login_instance)
            coordenadas = {
                'nome': dfCoordenadas['_pageName'],
                'urleviterbo': URL_EVITERBO + dfCoordenadas['_pageName'].replace(' ', '_', regex=True),
                'coord_1': dfCoordenadas['1'],
                'coord_5': dfCoordenadas['5'],
                'item_id': item_id
            }
            #wd_item = wdi_core.WDItemEngine(data=data) # create new item
            #wd_item.entity_metadata.get('id') #retriev item id

        print("===== DATAFRAME COORDENADAS FINAL ======")
        print(dfCoordenadas.values)
        dfWikibaseItems = pd.DataFrame(coordenadas,columns=['item_id'
            ,'nome'
            ,'urleviterbo'
            ,'coord_1'
            ,'coord_5'])

        print(dfWikibaseItems)
        dfWikibaseItems.to_csv(datasetpath+'coordenadas_3_wikibase_item_Teste.csv', index = True, header=True)

    '''=== POPULATE POSTO PROPERTIES ENTITY ==='''
    def create_posto(self, ws):
        posto_entity_values = {}
        posto_props = {}
        for row in ws.iter_cols(min_col=3,
                                max_col=3,
                                min_row=2,
                                max_row=12,
                                values_only=True):
            posto_prop_id = 'P50'
            posto_entity_values = {
                'P55':'',
                'P13': row,
                'P14': row}
        posto_props[posto_prop_id]=posto_entity_values
        print(posto_props)
        print(posto_entity_values)
        for key in posto_entity_values.keys():
            prop_id = key
        print(prop_id)
        list = posto_entity_values['P13']
        for i in range(len(list)):
            print(list[i])
            data=[wdi_core.WDUrl(prop_nr='P55', value=None, snak_type='somevalue')
                ,wdi_core.WDString(prop_nr='P13', value=list[i])
                ,wdi_core.WDString(prop_nr='P14', value=list[i])]
            #wd_item = wdi_core.WDItemEngine(wd_item_id=list[i],data=data)
            #wd_item.set_label(value[i], lang="pt")
            wd_item = wdi_core.WDItemEngine(data=data)
            wd_item.set_label(list[i], lang="pt")
            wd_item.write(self.login_instance)

    '''=== POPULATE ARMAS PROPERTIES ENTITY ==='''
    def create_armas(self, ws):
        armas_entity_values = {}
        armas_props = {}
        for row in ws.iter_cols(min_col=4,
                                max_col=4,
                                min_row=2,
                                max_row=4,
                                values_only=True):
            armas_prop_id = 'P52'
            armas_entity_values = {
                'P55':'',
                'P13': row,
                'P14': row,
                'P53': row
            }
        armas_props[armas_prop_id]=armas_entity_values
        print(armas_props)
        print(armas_entity_values)
        for key in armas_entity_values.keys():
            prop_id = key
        print(prop_id)
        list = armas_entity_values['P13']
        for i in range(len(list)):
            print(list[i])
            data=[wdi_core.WDUrl(prop_nr='P55', value=None, snak_type='somevalue')
                ,wdi_core.WDString(prop_nr='P13', value=list[i])
                ,wdi_core.WDString(prop_nr='P14', value=list[i])
                ,wdi_core.WDString(prop_nr='P53', value=list[i])]
            #wd_item = wdi_core.WDItemEngine(wd_item_id=list[i],data=data)
            #wd_item.set_label(value[i], lang="pt")
            wd_item = wdi_core.WDItemEngine(data=data)
            wd_item.set_label(list[i], lang="pt")
            wd_item.write(self.login_instance)


    '''=== POPULATE CARGO PROPERTIES ENTITY ==='''
    def create_cargos(self, ws):
        cargo_entity_values = {}
        cargo_props = {}
        for row in ws.iter_cols(min_col=5,
                                max_col=5,
                                min_row=2,
                                max_row=23,
                                values_only=True):
            cargo_prop_id = 'P8'
            cargo_entity_values = {
                'P9': '',
                'P10': '',
                'P13': row,
                'P14': '',
                'P16': ''
            }
        cargo_props[cargo_prop_id]=cargo_entity_values
        P16 = "https://schema.org/Occupation"
        print(cargo_props)
        print(cargo_entity_values)
        for key in cargo_entity_values.keys():
            prop_id = key
        print(prop_id)
        list = cargo_entity_values['P13']
        for i in range(len(list)):
            print(list[i])
            data=[wdi_core.WDString(prop_nr='P13', value=list[i])
                , wdi_core.WDString(prop_nr='P14', value=None, snak_type='somevalue')
                , wdi_core.WDUrl(prop_nr='P16', value=P16)
                , wdi_core.WDTime(prop_nr='P9', time=None, snak_type='somevalue')
                , wdi_core.WDTime(prop_nr='P10', time=None, snak_type='somevalue')]
            #wd_item = wdi_core.WDItemEngine(wd_item_id=list[i],data=data)
            #wd_item.set_label(list[i], lang="pt")
            wd_item = wdi_core.WDItemEngine(data=data)
            wd_item.set_label(list[i], lang="pt")
            wd_item.write(self.login_instance)

    '''=== POPULATE ACTIVIDADE PROPERTIES ENTITY ==='''
    def create_actividade(self, ws):
        actividade_entity_values = {}
        actividade_props = {}
        for row in ws.iter_cols(min_col=6,
                                max_col=6,
                                min_row=2,
                                max_row=20,
                                values_only=True):
            actividade_prop_id = 'P56'
            actividade_entity_values = {
                'P30':'',
                'P13': row,
                'P14': row
            }
        actividade_props[actividade_prop_id]=actividade_entity_values
        print(actividade_props)
        print(actividade_entity_values)
        for key in actividade_entity_values.keys():
            prop_id = key
        print(prop_id)
        list = actividade_entity_values['P13']
        for i in range(len(list)):
            print(list[i])
            data=[wdi_core.WDUrl(prop_nr='P30', value=None, snak_type='somevalue')
                ,wdi_core.WDString(prop_nr='P13', value=list[i])
                ,wdi_core.WDString(prop_nr='P14', value=None, snak_type='somevalue')]
            #wd_item = wdi_core.WDItemEngine(wd_item_id=list[i],data=data)
            #wd_item.set_label(value[i], lang="pt")
            wd_item = wdi_core.WDItemEngine(data=data)
            wd_item.set_label(list[i], lang="pt")
            wd_item.write(self.login_instance)

    def create_new_wikibase_item(self,data):
        wd_item = wdi_core.WDItemEngine(data=data)
        wd_item.write(self.login_instance)

    def create_missing_item(self, df_pessoas, i):
        nome_outras_grafias = self.get_Properties_Values(df_pessoas['nome_outras_grafias'][i])
        local_nascimento = self.get_Wikibase_Qid(df_pessoas['local_nascimento'][i])
        local_morte = self.get_Wikibase_Qid(df_pessoas['local_morte'][i])
        if df_pessoas['religião'][i] not in religioes.keys():
            religiao_values = object_without_properties
        else: religiao_values = {'value': religioes[df_pessoas['religião'][i]], 'snak_type':'value'}
        nome_completo = object_without_properties
        if df_pessoas['nome_completo'][i] != 0:
            nome_completo = {'value': df_pessoas['nome_completo'][i],'snak_type':'value' }
        residencias = []
        qualifiers_residencia = []
        qualifiers_formacao = []
        qualifiers_cargo = []
        qualifiers_posto = []
        qualifiers_actividade = []
        data= [wdi_core.WDString(prop_nr='P34', value=df_pessoas['_pageName'][i], snak_type='value'),
               wdi_core.WDString(prop_nr='P17', value=nome_completo['value'], snak_type=nome_completo['snak_type']),
               wdi_core.WDString(prop_nr='P20', value=nome_outras_grafias['value'],
                                 snak_type=nome_outras_grafias['snak_type'])
               ]
        if df_pessoas['pai'][i] != 0:
            description_pai = 'Esse objeto representa o pai da pessoa: ' + df_pessoas['_pageName'][i]
            QID_pai = self.create_new_item_on_wiki(df_pessoas['pai'][i], description_pai)
            data.append(wdi_core.WDItemID(prop_nr='P22', value=QID_pai,snak_type='value'))
        else: data.append(wdi_core.WDItemID(prop_nr='P22', value=object_without_properties['value']
                                            ,snak_type=object_without_properties['snak_type']))
        if df_pessoas['mãe'][i] != 0:
            description_mae = 'Esse objeto representa a mãe da pessoa: ' + df_pessoas['_pageName'][i]
            QID_mae = self.create_new_item_on_wiki(df_pessoas['mãe'][i], description_mae)
            data.append(wdi_core.WDItemID(prop_nr='P23', value=QID_mae,snak_type='value'))
        else: data.append(wdi_core.WDItemID(prop_nr='P23', value=object_without_properties['value']
                                            ,snak_type=object_without_properties['snak_type']))
        print("inserting brothers: ")
        print(df_pessoas['irmãos'][i])
        if df_pessoas['irmãos'][i] != 0:
            description_irmaos = 'Esse objeto representa o irmão da pessoa: ' + df_pessoas['_pageName'][i]
            QID_irmaos = self.create_new_item_on_wiki(df_pessoas['irmãos'][i], description_irmaos)
            data.append(wdi_core.WDItemID(prop_nr='P25', value=QID_irmaos,snak_type='value'))
        else: data.append(wdi_core.WDItemID(prop_nr='P25', value=object_without_properties['value']
                                            ,snak_type=object_without_properties['snak_type']))
        if df_pessoas['cônjuge'][i] != 0:
            description_conjuge = 'Esse objeto representa o cônjuge da pessoa: ' + df_pessoas['_pageName'][i]
            QID_conjuge = self.create_new_item_on_wiki(df_pessoas['cônjuge'][i], description_conjuge)
            data.append(wdi_core.WDItemID(prop_nr='P24', value=QID_conjuge,snak_type='value'))
        else: data.append(wdi_core.WDItemID(prop_nr='P24', value=object_without_properties['value']
                                            ,snak_type=object_without_properties['snak_type']))
        if df_pessoas['filhos'][i] != 0:
            description_filhos = 'Esse objeto representa o filho da pessoa: ' + df_pessoas['_pageName'][i]
            QID_filho = self.create_new_item_on_wiki(df_pessoas['filhos'][i], description_filhos)
            data.append(wdi_core.WDItemID(prop_nr='P21', value=QID_filho,snak_type='value'))
        else: data.append(wdi_core.WDItemID(prop_nr='P21', value=object_without_properties['value']
                                            ,snak_type=object_without_properties['snak_type']))
        data_nascimento = self.adjust_Time_Value(df_pessoas['data_nascimento'][i])
        if data_nascimento != '0':
            data.append(wdi_core.WDTime(prop_nr='P2', time=self.parse_To_Wikibase_Time(data_nascimento)))
        else: data.append(wdi_core.WDTime(prop_nr='P2', time=None, snak_type='somevalue'))
        data.append(wdi_core.WDItemID(prop_nr='P4', value=local_nascimento['value'],
                                      snak_type=local_nascimento['snak_type']))
        data_morte = self.adjust_Time_Value(df_pessoas['data_morte'][i])
        if data_morte != '0':
            data.append(wdi_core.WDTime(prop_nr='P3', time=self.parse_To_Wikibase_Time(data_morte)))
        else: data.append(wdi_core.WDTime(prop_nr='P3', time=None, snak_type='somevalue'))
        data.append(wdi_core.WDItemID(prop_nr='P5', value=local_morte['value'], snak_type=local_morte['snak_type']))
        if df_pessoas['sexo'][i] == 'masculino':
            data.append(wdi_core.WDItemID(prop_nr='P6', value='Q6',snak_type='value'))
        elif df_pessoas['sexo'][i] == 'feminino':
            data.append(wdi_core.WDItemID(prop_nr='P6', value='Q6',snak_type='value'))
        else: data.append(wdi_core.WDItemID(prop_nr='P6', value=object_without_properties['value']
                                            ,snak_type=object_without_properties['snak_type']))
        data.append(wdi_core.WDItemID(prop_nr='P27', value=religiao_values['value'],
                                      snak_type=religiao_values['snak_type']))
        for idx_res in range(len(column_residencia)):
            idx_dt_ini = idx_res + 1
            residencias.append(self.get_Wikibase_Qid(df_pessoas[column_residencia[idx_res]][i]))
            resid_data_ini = self.adjust_Time_Value(df_pessoas[datas_qualifiers['residência'][0]+ str(idx_dt_ini)][i])
            resid_data_fim = self.adjust_Time_Value(df_pessoas[datas_qualifiers['residência'][1]+ str(idx_dt_ini)][i])
            if resid_data_ini != '0':
                qualifiers_residencia.append(wdi_core.WDTime(prop_nr='P9',
                                                             time=self.parse_To_Wikibase_Time(resid_data_ini),
                                                             is_qualifier=True))
            else: qualifiers_residencia.append(wdi_core.WDTime(prop_nr='P9',
                                                               time=None, snak_type='somevalue',is_qualifier=True))
            if resid_data_fim != '0':
                qualifiers_residencia.append(wdi_core.WDTime(prop_nr='P10',
                                                             time=self.parse_To_Wikibase_Time(resid_data_fim),
                                                             is_qualifier=True))
            else: qualifiers_residencia.append(wdi_core.WDTime(prop_nr='P10',
                                                               time=None, snak_type='somevalue',is_qualifier=True))
        instituicoes_de_formacoes = [self.get_Wikibase_Qid(df_pessoas['Instituição_de_Formação_1'][i])
            ,self.get_Wikibase_Qid(df_pessoas['Instituição_de_Formação_2'][i])
            ,self.get_Wikibase_Qid(df_pessoas['Instituição_de_Formação_3'][i])
                                     ]
        formacoes = []
        for idx_f in range(len(column_formacao)):
            idx_formacao = idx_f + 1
            formacoes.append(self.get_Wikibase_Qid(df_pessoas[column_formacao[idx_f]][i]))
            formacao_dt_ini = self.adjust_Time_Value(df_pessoas[datas_qualifiers['formação'][0] + str(idx_formacao)][i])
            formacao_dt_fim = self.adjust_Time_Value(df_pessoas[datas_qualifiers['formação'][1] + str(idx_formacao)][i])
            if formacao_dt_ini != '0':
                qualifiers_formacao.append(wdi_core.WDTime(prop_nr='P9',
                                                           time=self.parse_To_Wikibase_Time(formacao_dt_ini),
                                                           is_qualifier=True))
            else: qualifiers_formacao.append(wdi_core.WDTime(prop_nr='P9',time=None, snak_type='somevalue',is_qualifier=True))
            if formacao_dt_fim != '0':
                qualifiers_formacao.append(wdi_core.WDTime(prop_nr='P10',
                                                           time=self.parse_To_Wikibase_Time(formacao_dt_fim),
                                                           is_qualifier=True))
            else: qualifiers_formacao.append(wdi_core.WDTime(prop_nr='P10',time=None, snak_type='somevalue',is_qualifier=True))

            qualifiers_formacao.append(wdi_core.WDItemID(prop_nr='P28', value=instituicoes_de_formacoes[idx_f]['value'],
                                                         snak_type=instituicoes_de_formacoes[idx_f]['snak_type'],
                                                         is_qualifier=True))
            qualifiers_formacao.append(wdi_core.WDItemID(prop_nr='P12', value=instituicoes_de_formacoes[idx_f]['value'],
                                                         snak_type=instituicoes_de_formacoes[idx_f]['snak_type'],
                                                         is_qualifier=True))
        postos = []
        armas = []
        for idx_p in range(len(column_posto)):
            idx_posto = idx_p + 1
            print("appending " + str(column_posto[idx_p]) + " : " + str(df_pessoas[column_posto[idx_p]][i]) + " of position " + str(i))
            postos.append(self.get_Wikibase_Qid(df_pessoas[column_posto[idx_p]][i]))
            posto_dt_ini = self.adjust_Time_Value(df_pessoas[datas_qualifiers['posto'][0]+ str(idx_posto)][i])
            posto_dt_fim = self.adjust_Time_Value(df_pessoas[datas_qualifiers['posto'][1]+ str(idx_posto)][i])
            if posto_dt_ini != '0':
                qualifiers_posto.append(wdi_core.WDTime(prop_nr='P9',
                                                        time=self.parse_To_Wikibase_Time(posto_dt_ini),
                                                        is_qualifier=True))
            else:qualifiers_posto.append(wdi_core.WDTime(prop_nr='P9',time=None,snak_type='somevalue',is_qualifier=True))
            if posto_dt_fim != '0':
                qualifiers_posto.append(wdi_core.WDTime(prop_nr='P10',
                                                        time=self.parse_To_Wikibase_Time(posto_dt_fim),
                                                        is_qualifier=True))
            else:qualifiers_posto.append(wdi_core.WDTime(prop_nr='P10',time=None,snak_type='somevalue',is_qualifier=True))
            armas.append(self.get_Wikibase_Qid(df_pessoas[column_arma[idx_p]][i]))
            qualifiers_posto.append(wdi_core.WDItemID(prop_nr='P52', value=armas[idx_p]['value'],
                                                      snak_type=armas[idx_p]['snak_type'],
                                                      is_qualifier=True))
        cargos = []
        instituicoes_cargos = []
        for idx_c in range(len(column_cargo)):
            idx_cargo = idx_c + 1
            cargos.append(self.get_Wikibase_Qid(df_pessoas[column_cargo[idx_c]][i]))
            cargo_dt_ini = self.adjust_Time_Value(df_pessoas[datas_qualifiers['cargo'][0]+ str(idx_cargo)][i])
            cargo_dt_fim = self.adjust_Time_Value(df_pessoas[datas_qualifiers['cargo'][1]+ str(idx_cargo)][i])
            if cargo_dt_ini != '0':
                qualifiers_cargo.append(wdi_core.WDTime(prop_nr='P9',
                                                        time=self.parse_To_Wikibase_Time(cargo_dt_ini),
                                                        is_qualifier=True))
            else: qualifiers_cargo.append(wdi_core.WDTime(prop_nr='P9',time=None, snak_type='somevalue',is_qualifier=True))
            if cargo_dt_fim != '0':
                qualifiers_cargo.append(wdi_core.WDTime(prop_nr='P10',
                                                        time=self.parse_To_Wikibase_Time(cargo_dt_fim),
                                                        is_qualifier=True))
            else:qualifiers_cargo.append(wdi_core.WDTime(prop_nr='P10',time=None, snak_type='somevalue',is_qualifier=True))
            instituicoes_cargos.append(self.get_Wikibase_Qid(df_pessoas[column_inst_cargo[idx_c]][i]))
            qualifiers_cargo.append(wdi_core.WDItemID(prop_nr='P12', value=instituicoes_cargos[idx_c]['value'],
                                                      snak_type=instituicoes_cargos[idx_c]['snak_type'],
                                                      is_qualifier=True))
        actividades = []
        for idx_ac in range(len(column_actividade)):
            idx_actividade = idx_ac + 1
            actividades.append(self.get_Wikibase_Qid(df_pessoas[column_actividade[idx_ac]][i]))
            actividade_dt_ini = self.adjust_Time_Value(df_pessoas[datas_qualifiers['actividade'][0]+ str(idx_actividade)][i])
            actividade_dt_fim = self.adjust_Time_Value(df_pessoas[datas_qualifiers['actividade'][1]+ str(idx_actividade)][i])
            if actividade_dt_ini != '0':
                qualifiers_actividade.append(wdi_core.WDTime(prop_nr='P9',
                                                             time=self.parse_To_Wikibase_Time(actividade_dt_ini),
                                                             is_qualifier=True))
            else: qualifiers_actividade.append(wdi_core.WDTime(prop_nr='P9',time=None,snak_type='somevalue',is_qualifier=True))
            if actividade_dt_fim != '0':
                qualifiers_actividade.append(wdi_core.WDTime(prop_nr='P10',time=self.parse_To_Wikibase_Time(actividade_dt_fim),
                                                             is_qualifier=True))
            else:qualifiers_actividade.append(wdi_core.WDTime(prop_nr='P10',time=None,snak_type='somevalue',is_qualifier=True))

        locais_de_actividades = []
        for idx_loc_ac in range(len(column_loc_actividade)):
            locais_de_actividades.append(self.get_Wikibase_Qid(df_pessoas[column_loc_actividade[idx_loc_ac]][i]))
            qualifiers_actividade.append(wdi_core.WDItemID(prop_nr='P57', value=locais_de_actividades[idx_loc_ac]['value'],
                                                           snak_type=locais_de_actividades[idx_loc_ac]['snak_type'],
                                                           is_qualifier=True))
        print('======== TESTAR INSERÇÃO QUALIFIERS RESIDENCIAS PARA UM REGISTRO!! ======')
        for idx_residencias in range(len(residencias)):
            data.append(wdi_core.WDItemID(prop_nr='P26', value=residencias[idx_residencias]['value'],
                                          snak_type=residencias[idx_residencias]['snak_type'],
                                          qualifiers=qualifiers_residencia))
        for idx_formacoes in range(len(formacoes)):
            data.append(wdi_core.WDItemID(prop_nr='P48', value=formacoes[idx_formacoes]['value'],
                                          snak_type=formacoes[idx_formacoes]['snak_type'],
                                          qualifiers=qualifiers_formacao))
        for idx_postos in range(len(postos)):
            print("inserting posto " + str(postos[idx_postos]['value']))
            data.append(wdi_core.WDItemID(prop_nr='P50', value=postos[idx_postos]['value'],
                                          snak_type=postos[idx_postos]['snak_type'],
                                          qualifiers=qualifiers_posto))
        for idx_cargos in range(len(cargos)):
            print("inserting posto " + str(cargos[idx_cargos]['value']))
            data.append(wdi_core.WDItemID(prop_nr='P8', value=cargos[idx_cargos]['value'],
                                          snak_type=cargos[idx_cargos]['snak_type'],
                                          qualifiers=qualifiers_cargo))
        for idx_actividades in range(len(actividades)):
            data.append(wdi_core.WDItemID(prop_nr='P56', value=actividades[idx_actividades]['value'],
                                          snak_type=actividades[idx_actividades]['snak_type'],
                                          qualifiers=qualifiers_actividade))
        url_eviterbo_pessoa = URL_EVITERBO + df_pessoas['_pageName'].replace(' ', '_',regex=True).to_string()
        url_eviterbo_pessoa = url_eviterbo_pessoa.replace('0 ','')
        url_eviterbo_pessoa = url_eviterbo_pessoa.strip()
        data.append(wdi_core.WDUrl(prop_nr='P30', value=None,snak_type='somevalue'))
        wd_item = wdi_core.WDItemEngine(data=data) # create new item
        wd_item.set_label(label=df_pessoas['_pageName'][i],lang='pt')
        wd_item.set_description(description='Pessoa',lang='pt')
        wd_item.write(self.login_instance)
        item_id = wd_item.entity_metadata.get('id') #retrieve item id

    def editWikiPages(self, infobox_bottom_text, df_pessoas):


        S = requests.Session()
        # Retrieve login token first
        PARAMS_0 = {
            'action':"query",
            'meta':"tokens",
            'type':"login",
            'format':"json"
        }
        R = S.get(url=URL, params=PARAMS_0)
        DATA = R.json()
        #print(DATA)
        LOGIN_TOKEN = DATA['query']['tokens']['logintoken']
        #print(LOGIN_TOKEN)

        # Send a post request to login. Using the main account for login is not
        # supported. Obtain credentials via Special:BotPasswords
        # (https://www.mediawiki.org/wiki/Special:BotPasswords) for lgname & lgpassword
        PARAMS_1 = {
            'action':"login",
            'lgname':"Operador@eviterboManut",
            'lgpassword':"1vt8m6bd8v1lsjo93u6v6l23ltl7shld",
            'lgtoken':LOGIN_TOKEN,
            'format':"json"
        }
        R = S.post(URL, data=PARAMS_1)
        DATA = R.json()
        #print(DATA)
        # Step 3: GET request to fetch CSRF token
        PARAMS_2 = {
            "action": "query",
            "meta": "tokens",
            "format": "json"
        }
        R = S.get(url=URL, params=PARAMS_2)
        DATA = R.json()
        CSRF_TOKEN = DATA['query']['tokens']['csrftoken']
        PARAMS_3={
            "action": "logout",
            "token": CSRF_TOKEN,
            "format": "json"
        }
        for k in range(len(df_pessoas)):
            SEARCHPAGE = df_pessoas['_pageName'][k]
            pageid = df_pessoas['_pageID'][k]
            PARAMS_4 = {
                "action": "parse",
                "pageid": pageid,
                "prop":"wikitext",
                "format": "json"
            }
            R = S.post(URL, data=PARAMS_4)
            DATA = R.json()
            #print(DATA)
            logging.debug(DATA["parse"]["wikitext"]["*"])
            infobox_text_to_replace = DATA["parse"]["wikitext"]["*"]
            final_index = "\n}}" #"\n\n\n==Biografia=="
            Q_ITEM = self.get_Wikibase_Qid(SEARCHPAGE)
            top_text = "{{Info/TechNetEMPIRE"
            varentity = "{{#vardefine:entity|"
            if Q_ITEM['value'] is not None:
                infobox_complete_text = f"{top_text}\n <!-- dados da imagem(s) -->\n" \
                                        f"{varentity}{Q_ITEM['value']}}}}}\n{infobox_bottom_text}"
                # EDITAR A PÁGINA
                logging.debug(infobox_complete_text)
                logging.debug("==== GETTING FINAL PART OF THE TEXT: \n")
                start_final_text = infobox_text_to_replace.find(final_index)
                start_final_text+=3
                final_text = infobox_text_to_replace[start_final_text:]
                wiki_final_text = infobox_complete_text
                logging.debug("MERGING THE STRING TO WRITE INTO WIKI PAGE: \n")
                wiki_final_text = wiki_final_text + final_text
                PARAMS_5 = {
                    "action": "edit",
                    "pageid": pageid,
                    "token": CSRF_TOKEN,
                    "format": "json",
                    "text": wiki_final_text
                }
                print("WRITING INTO WIKI PAGE ...")

                try:
                    R = S.post(URL, data=PARAMS_5, verify=False)
                    DATA = R.json()
                    logging.debug(DATA)
                except ConnectionError:
                    R = S.post(URL, data=PARAMS_3)
                    DATA = R.json()
                    print(DATA)

            else: self.create_missing_item(df_pessoas, k)
        R = S.post(URL, data=PARAMS_3)
        DATA = R.json()
        print(DATA)

    '''=== UPDATE INSTITUICOES === '''
    def update_instituicoes(self):
        columns_Instituicao = ['_pageName','outras_denominações','tipo','data_fundação','data_extinção'
            ,'paralisação_início','paralisação_fim','Localização1','Localização_data_início1'
            ,'Localização_data_fim1','Localização2','Localização_data_início2','Localização_data_fim2'
            ,'Localização3','Localização_data_início3','Localização_data_fim3','Localização4'
            ,'Localização_data_início4','Localização_data_fim4','Localização5','Localização_data_início5'
            ,'Localização_data_fim5','Localização6','Localização_data_início6','Localização_data_fim6','Localização7'
            ,'Localização_data_início7','Localização_data_fim7','Localização8','Localização_data_início8'
            ,'Localização_data_fim8','antecessora','sucessora']
        r_file = open(datasetpath+instituicao_file, 'rb').read()
        result = chardet.detect(r_file)
        file_encoding = result['encoding']
        dfInstituicao = pd.read_csv(datasetpath+instituicao_file
                                    ,encoding=file_encoding
                                    ,header=0
                                    ,usecols=columns_Instituicao)
        dfInstituicao = dfInstituicao.reindex(columns=dfInstituicao.columns.tolist())
        dfInstNames = dfInstituicao['_pageName'].replace(' ', '_', regex=True)
        dfInstituicao = dfInstituicao.replace(['\[','\]'], ['',''], regex=True)
        dfInstituicao = dfInstituicao.fillna(0)
        print(dfInstNames.head(10))
        for i in range(len(dfInstituicao)):
            inst_name = dfInstituicao['_pageName'][i]
            print("Updating institution {}  ...".format(inst_name))
            Q_instituicao = self.get_Wikibase_Qid(inst_name, 'Instituição')
            item_id_updated = self.add_properties_and_qualifiers(Q_instituicao, dfInstituicao, i,'P35', 'P9')
if __name__ == '__main__':

    writer = WikibaseWriter()
    #df_pessoas_vazias = writer.initiate()
    # for i in range(len(df_pessoas_vazias)):
    #     writer.create_missing_item(df_pessoas_vazias, i)
    writer.update_instituicoes()
    with open('data.txt', 'r') as file:
        infobox_complement_text = file.read()#.replace("\n","")

    file.close()
    #writer.editWikiPages(infobox_complement_text,writer.initiate())
