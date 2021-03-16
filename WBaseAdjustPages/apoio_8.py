
import json
import openpyxl
import requests
from mwclient import Site
from wikibase_api import Wikibase
#import wikidataintegrator
from python_wikibase import PyWikibase
import chardet
import pandas as pd
# import pprint
#
from wikidataintegrator.wdi_config import config
from wikidataintegrator import wdi_login, wdi_core

PATH_EXCEL_PROPERTIES = "/../WBaseAdjustPages/Propriedades.xlsx"
datasetpath = "/../WBaseAdjustPages/"
coord_file = "coordenadasWikibase-May2020.csv"
instituicao_file = "instituicaoWikibase-May2020.csv"
pessoa_file = "pessoaWikibase-May2020.csv"
URL = "http://ieeta-eviterbo.web.ua.pt/api.php"
URL_EVITERBO = "http://ieeta-eviterbo.web.ua.pt/index.php/"
# WikidataIntegrator Configuration
config['MEDIAWIKI_API_URL'] = "http://ieeta-eviterbo.web.ua.pt/api.php"
config['WIKIBASE_URL'] = 'http://ieeta-eviterbo.web.ua.pt'
config['CONCEPT_BASE_URI'] = 'http://ieeta-eviterbo.web.ua.pt/entity/'
config['CALENDAR_MODEL_QID'] = 'http://www.wikidata.org/entity/Q1985727'

class WikibaseWriter(object):
    def __init__(self):
        self.login_credentials = {
            "bot_username": "WikiBase",
            "bot_password": "passwd",
        }
        self.login_instance = wdi_login.WDLogin(user= "WikiBase", pwd="passwd")
        excelfile = openpyxl.load_workbook(PATH_EXCEL_PROPERTIES)
        self.ws = excelfile.active
        self.maxRows = self.ws.max_row
        self.maxCol = self.ws.max_column

    def initiate(self):
        dfWikibaseItems = pd.DataFrame
        #wd_item = wdi_core.WDItemEngine(data=data) # create new item
        #wd_item.entity_metadata.get('id') #retriev item id
        columns_inst = [1,7,8,9,19,11,12,13,14,15,16,17,18,19,20,21,22,23,24,25,26,27,28,29,30,31,32
                        ,33,34,35,36,37]
        columns_Instituicao = ['_pageName','tipo','data_fundação','data_extinção'
            ,'paralisação_início','paralisação_fim','Localização1','Localização_data_início1'
            ,'Localização_data_fim1','Localização2','Localização_data_início2','Localização_data_fim2'
            ,'Localização3','Localização_data_início3','Localização_data_fim3','Localização4'
            ,'Localização_data_início4','Localização_data_fim4','Localização5','Localização_data_início5'
            ,'Localização_data_fim5','Localização6','Localização_data_início6','Localização_data_fim6','Localização7'
            ,'Localização_data_início7','Localização_data_fim7','Localização8','Localização_data_início8'
            ,'Localização_data_fim8','antecessora','sucessora']
        columns_Coordenada = ['_pageName','1','5']
        r_file = open(datasetpath+coord_file, 'rb').read()
        result = chardet.detect(r_file)
        file_encoding = result['encoding']
        dfInstituicao = pd.read_csv(datasetpath+instituicao_file
                                    ,encoding=file_encoding
                                    ,header=0
                                    ,usecols=columns_inst)

        dfCoordenadas = pd.read_csv(datasetpath+coord_file
                                    ,encoding=file_encoding
                                    ,header=0
                                    ,nrows=175
                                    ,usecols=columns_Coordenada)

        dfCoordenadas_Encode = pd.read_csv(datasetpath+coord_file
                                           ,encoding=file_encoding
                                           ,header=0
                                           ,skiprows=175)
                                           #,usecols=columns_Coordenada)

        dfCoordenadas = dfCoordenadas.reindex(columns=dfCoordenadas.columns.tolist())
        dataCoordNames = dfCoordenadas['_pageName'].replace(' ', '_', regex=True)
        #dataTemp = dataTemp.replace(['\[','\]'], ['',''], regex=True)
        coordenadas = {}
        #print(dataCoordNames.head(10))
        #frame = pd.DataFrame(dataTemp['_pageName'])
        #print(dataCoordNames.values)

        for i in range(len(dfCoordenadas)):
            coord_name = dfCoordenadas['_pageName']
            url_coord=URL_EVITERBO + dfCoordenadas['_pageName'].replace(' ', '_', regex=True)
            lat = dfCoordenadas['1']
            long = dfCoordenadas['5']
            data=[wdi_core.WDUrl(prop_nr='P30',value=url_coord)
                 ,wdi_core.WDGlobeCoordinate(prop_nr='P29', latitude=lat,longitude=long,precision=0.0001)
                 ,wdi_core.WDString(prop_nr='P45', value=dfCoordenadas['_pageName'])]
            # wd_item = wdi_core.WDItemEngine(data=data) # create new item
            # wd_item.set_label(label=dfCoordenadas['_pageName'],lang='pt')
            item_id = 'Q001'#wd_item.entity_metadata.get('id') #retrieve item id
            # wd_item.write(self.login_instance)
            coordenadas = {
                'nome': dfCoordenadas['_pageName'],
                'urleviterbo': URL_EVITERBO + dfCoordenadas['_pageName'].replace(' ', '_', regex=True),
                'coord_1': dfCoordenadas['1'],
                'coord_5': dfCoordenadas['5'],
                'item_id': item_id
            }
            #wd_item = wdi_core.WDItemEngine(data=data) # create new item
            #wd_item.entity_metadata.get('id') #retriev item id